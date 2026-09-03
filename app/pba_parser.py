"""Парсер ПБА-файлов Gedeon Richter (xlsx) в структуру для отчёта.

Структура файла (одинаковая у weekly- и monthly-выгрузок AMK):
    Клиент: | ТОО «Gedeon Richter KZ»
    Бренд:  | Стопдиар
    Период: | 2026-08-01
    тотал                          <- блок: месяц целиком по площадкам
    Платформа | Формат | Формат закупки | KPI (план/факт/разница/% от плана) | Бюджет с НДС и АК (план/факт/разница)
    ...строки...
    Всего     <- итог блока
    1 неделя | 25.07 - 02.08       <- блок недели (date_range во второй колонке, есть не всегда)
    ...
Один xlsx = несколько листов (месяцев), каждый лист парсится независимо.

ВАЖНО про честность отчёта: у недель, которые ещё не наступили или не завершены,
факт = 0 / пусто. Считать «% от плана» по месяцу целиком нельзя — получится ложное
недоосвоение. Поэтому недели делятся на закрытые (есть факт бюджета) и незакрытые,
и все выводы строятся по закрытым.
"""

from __future__ import annotations

import re
from datetime import date, datetime

import openpyxl

WEEK_RE = re.compile(r"^\s*(\d+)\s*недел[яию]\s*$", re.IGNORECASE)
TOTAL_BLOCK_RE = re.compile(r"^\s*тотал\s*$", re.IGNORECASE)
BLOCK_END_RE = re.compile(r"^\s*всего\s*$", re.IGNORECASE)

# Колонки блока (0-based), как в выгрузке AMK.
COL_PLATFORM, COL_FORMAT, COL_BUY_MODEL = 0, 1, 2
COL_KPI_PLAN, COL_KPI_FACT, COL_KPI_DIFF, COL_KPI_PCT = 3, 4, 5, 6
COL_BUDGET_PLAN, COL_BUDGET_FACT = 7, 8

# Потолки на размер файла: 20 МБ xlsx распаковывается в гигабайты, а лист на миллионы
# ячеек съест память процесса. Реальный ПБА — это единицы листов и десятки строк,
# так что запас тут огромный.
MAX_SHEETS = 24
MAX_ROWS_PER_SHEET = 4000


class ParseError(ValueError):
    """Файл не похож на ПБА — сообщение показывается пользователю как есть."""


def _num(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _row_dict(row) -> dict:
    kpi_plan = _num(row[COL_KPI_PLAN])
    kpi_fact = _num(row[COL_KPI_FACT])
    budget_plan = _num(row[COL_BUDGET_PLAN])
    budget_fact = _num(row[COL_BUDGET_FACT])
    row_dict = {
        "platform": _text(row[COL_PLATFORM]),
        "format": _text(row[COL_FORMAT]),
        "buy_model": _text(row[COL_BUY_MODEL]),
        "kpi_plan": kpi_plan,
        "kpi_fact": kpi_fact,
        "kpi_pct": _num(row[COL_KPI_PCT]),
        "budget_plan": budget_plan,
        "budget_fact": budget_fact,
        "budget_pct": (budget_fact - budget_plan) / budget_plan
        if budget_plan and budget_fact is not None
        else None,
        "unit_cost_plan": _unit_cost(_text(row[COL_BUY_MODEL]), budget_plan, kpi_plan),
        "unit_cost_fact": _unit_cost(_text(row[COL_BUY_MODEL]), budget_fact, kpi_fact),
    }
    row_dict["unit_cost_pct"] = (
        (row_dict["unit_cost_fact"] - row_dict["unit_cost_plan"]) / row_dict["unit_cost_plan"]
        if row_dict["unit_cost_plan"] and row_dict["unit_cost_fact"] is not None
        else None
    )
    return row_dict


def _parse_meta(rows: list[tuple]) -> dict:
    meta = {"client": "", "brand": "", "period": ""}
    keys = (("клиент", "client"), ("бренд", "brand"), ("период", "period"))
    for row in rows[:8]:
        label = _text(row[0]).lower().rstrip(":")
        for needle, key in keys:
            if label == needle:
                meta[key] = _text(row[1])
    return meta


def _parse_blocks(rows: list[tuple]) -> list[dict]:
    blocks: list[dict] = []
    i = 0
    while i < len(rows):
        label = _text(rows[i][COL_PLATFORM])
        is_total = bool(TOTAL_BLOCK_RE.match(label))
        week_match = WEEK_RE.match(label)
        if not (is_total or week_match):
            i += 1
            continue

        date_range = _text(rows[i][1]) or None
        i += 3  # строка-заголовок блока + две строки шапки таблицы
        data_rows, total_row = [], None
        while i < len(rows):
            first = _text(rows[i][COL_PLATFORM])
            if BLOCK_END_RE.match(first):
                total_row = rows[i]
                break
            if first:
                row = _row_dict(rows[i])
                # Ни плана, ни факта — площадки в этом периоде не было. В отчёте такая
                # строка была бы шумом из нулей, а не информацией.
                if any(
                    (row["kpi_plan"], row["kpi_fact"], row["budget_plan"], row["budget_fact"])
                ):
                    data_rows.append(row)
            i += 1

        if total_row is None:
            continue

        total = _row_dict(total_row)
        blocks.append(
            {
                "kind": "month_total" if is_total else "week",
                "label": "Месяц целиком" if is_total else label,
                "week_number": None if is_total else int(week_match.group(1)),
                "date_range": date_range,
                "rows": data_rows,
                "total": total,
            }
        )
    return blocks


def _aggregate_closed(weeks: list[dict]) -> list[dict]:
    """Суммирует строки закрытых недель по (площадка, формат, модель закупки).

    KPI суммируется только внутри одной модели закупки — просмотры (CPV), показы (CPM)
    и клики (CPC) складывать между собой нельзя, получится бессмысленное число.
    """
    buckets: dict[tuple, dict] = {}
    for week in weeks:
        for row in week["rows"]:
            key = (row["platform"], row["format"], row["buy_model"])
            agg = buckets.setdefault(
                key,
                {
                    "platform": row["platform"],
                    "format": row["format"],
                    "buy_model": row["buy_model"],
                    "kpi_plan": 0.0,
                    "kpi_fact": 0.0,
                    "budget_plan": 0.0,
                    "budget_fact": 0.0,
                },
            )
            for field in ("kpi_plan", "kpi_fact", "budget_plan", "budget_fact"):
                agg[field] += row[field] or 0.0

    result = []
    for agg in buckets.values():
        agg["kpi_pct"] = (
            (agg["kpi_fact"] - agg["kpi_plan"]) / agg["kpi_plan"] if agg["kpi_plan"] else None
        )
        agg["budget_pct"] = (
            (agg["budget_fact"] - agg["budget_plan"]) / agg["budget_plan"]
            if agg["budget_plan"]
            else None
        )
        agg["unit_cost_plan"] = _unit_cost(agg["buy_model"], agg["budget_plan"], agg["kpi_plan"])
        agg["unit_cost_fact"] = _unit_cost(agg["buy_model"], agg["budget_fact"], agg["kpi_fact"])
        agg["unit_cost_pct"] = (
            (agg["unit_cost_fact"] - agg["unit_cost_plan"]) / agg["unit_cost_plan"]
            if agg["unit_cost_plan"] and agg["unit_cost_fact"] is not None
            else None
        )
        result.append(agg)
    result.sort(key=lambda a: -a["budget_plan"])
    return result


def _unit_cost(buy_model: str, budget: float | None, kpi: float | None) -> float | None:
    """Эффективная цена единицы: бюджет с НДС и АК / объём. Для CPM — за 1000."""
    if not budget or not kpi:
        return None
    model = (buy_model or "").upper()
    if "CPM" in model:
        return budget / kpi * 1000
    if "CPV" in model or "CPC" in model or "CPL" in model:
        return budget / kpi
    return None


def _parse_sheet(name: str, rows: list[tuple]) -> dict | None:
    blocks = _parse_blocks(rows)
    if not blocks:
        return None

    weeks = [b for b in blocks if b["kind"] == "week"]
    weeks.sort(key=lambda b: b["week_number"])
    month_total = next((b for b in blocks if b["kind"] == "month_total"), None)

    closed = [w for w in weeks if (w["total"]["budget_fact"] or 0) > 0]
    pending = [w for w in weeks if (w["total"]["budget_fact"] or 0) <= 0]

    plan_closed = sum(w["total"]["budget_plan"] or 0 for w in closed)
    fact_closed = sum(w["total"]["budget_fact"] or 0 for w in closed)
    plan_pending = sum(w["total"]["budget_plan"] or 0 for w in pending)
    plan_month = (
        month_total["total"]["budget_plan"]
        if month_total
        else sum(w["total"]["budget_plan"] or 0 for w in weeks)
    )

    return {
        "sheet": name,
        "label": name.strip(),
        "meta": _parse_meta(rows),
        "month_total": month_total,
        "weeks": weeks,
        "closed_week_numbers": [w["week_number"] for w in closed],
        "pending_week_numbers": [w["week_number"] for w in pending],
        "summary": {
            "plan_month": plan_month,
            "plan_closed": plan_closed,
            "fact_closed": fact_closed,
            "plan_pending": plan_pending,
            "delivery_pct": (fact_closed / plan_closed) if plan_closed else None,
            "closed_weeks": len(closed),
            "total_weeks": len(weeks),
        },
        "by_placement_closed": _aggregate_closed(closed),
    }


def parse_pba_file(path) -> dict:
    """Возвращает {'client', 'brand', 'months': [...]} — по одному месяцу на лист."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if len(wb.sheetnames) > MAX_SHEETS:
            raise ParseError(
                f"В файле {len(wb.sheetnames)} листов — это не похоже на ПБА "
                f"(ожидается до {MAX_SHEETS})."
            )
        months = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(tuple(row) + (None,) * max(0, 10 - len(row)))
                if len(rows) > MAX_ROWS_PER_SHEET:
                    raise ParseError(
                        f"На листе «{name}» больше {MAX_ROWS_PER_SHEET} строк — "
                        "это не похоже на ПБА."
                    )
            parsed = _parse_sheet(name, rows)
            if parsed:
                months.append(parsed)
    finally:
        wb.close()

    if not months:
        raise ParseError(
            "В файле не нашлось ни одного листа с ПБА: ожидались блоки «тотал» и «N неделя» "
            "с колонками Платформа / Формат / Формат закупки / KPI / Бюджет."
        )

    client = next((m["meta"]["client"] for m in months if m["meta"]["client"]), "")
    brand = next((m["meta"]["brand"] for m in months if m["meta"]["brand"]), "")
    return {"client": client, "brand": brand, "months": months}
